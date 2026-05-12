"""One-off / repeatable: convert FCB Excel sheets to ml-model-service JSON seeds."""
import json
import os
from datetime import date, datetime, timezone

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUT_DIR = os.path.join(BASE, "ml-model-service", "src", "main", "resources", "data")


def season_from_date(d: date) -> str:
    if d.month >= 7:
        return f"{d.year}-{str(d.year + 1)[-2:]}"
    return f"{d.year - 1}-{str(d.year)[-2:]}"


def parse_date(cell) -> date:
    if isinstance(cell, datetime):
        return cell.date()
    if isinstance(cell, date):
        return cell
    if isinstance(cell, str):
        return datetime.strptime(cell[:10], "%Y-%m-%d").date()
    raise ValueError(f"Unsupported date cell: {cell!r}")


def build_matches():
    path = os.path.join(BASE, "FCB_Football_Matches.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows[1:]:
        if not r or r[idx["Match_ID"]] is None:
            continue
        mid = str(r[idx["Match_ID"]]).strip()
        d = parse_date(r[idx["Date"]])
        comp = r[idx["Competition"]]
        venue = str(r[idx["Venue"]]).strip() if r[idx["Venue"]] else ""
        opp = str(r[idx["Opponent"]]).strip() if r[idx["Opponent"]] else ""
        bg = int(r[idx["Barca_Goals"]]) if r[idx["Barca_Goals"]] is not None else None
        og = int(r[idx["Opp_Goals"]]) if r[idx["Opp_Goals"]] is not None else None
        hxg = float(r[idx["Barca_xG"]]) if r[idx["Barca_xG"]] is not None else None
        oxg = float(r[idx["Opp_xG"]]) if r[idx["Opp_xG"]] is not None else None

        if venue.lower() == "home":
            home_team, away_team = "FC Barcelona", opp
            home_goals, away_goals = bg, og
            xg_home, xg_away = hxg, oxg
        else:
            home_team, away_team = opp, "FC Barcelona"
            home_goals, away_goals = og, bg
            xg_home, xg_away = oxg, hxg

        extra = {}
        skip = {
            "Match_ID",
            "Date",
            "Competition",
            "Venue",
            "Opponent",
            "Barca_Goals",
            "Opp_Goals",
            "Barca_xG",
            "Opp_xG",
        }
        for col in hdr:
            if col in skip:
                continue
            v = r[idx[col]]
            if v is None:
                continue
            key = col.replace("%", "_pct").replace(" ", "_")
            if isinstance(v, (int, float, str, bool)):
                extra[key] = v
            else:
                extra[key] = str(v)

        ts = datetime(d.year, d.month, d.day, 12, 0, 0, tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")

        rec = {
            "matchExternalKey": mid,
            "season": season_from_date(d),
            "matchDate": d.isoformat(),
            "homeTeamName": home_team,
            "awayTeamName": away_team,
            "homeGoals": home_goals,
            "awayGoals": away_goals,
            "competition": str(comp) if comp else None,
            "stadium": venue,
            "xgHome": xg_home,
            "xgAway": xg_away,
            "eventTimestamp": ts,
        }
        if extra:
            rec["extraFeatures"] = extra
        out.append(rec)
    wb.close()
    return out


def build_players():
    path = os.path.join(BASE, "Football_Players_Data.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    entity_cols = {
        "Player_ID",
        "Name",
        "Age",
        "Position",
        "Nationality",
        "Matches",
        "Goals",
        "Assists",
        "Market_Value_EUR",
    }
    out = []
    for r in rows[1:]:
        if not r or r[idx["Player_ID"]] is None:
            continue
        pid = str(r[idx["Player_ID"]]).strip()
        name = str(r[idx["Name"]]).strip() if r[idx["Name"]] else None
        age = int(r[idx["Age"]]) if r[idx["Age"]] is not None else None
        pos = str(r[idx["Position"]]).strip() if r[idx["Position"]] else None
        nat = str(r[idx["Nationality"]]).strip() if r[idx["Nationality"]] else None
        apps = int(r[idx["Matches"]]) if r[idx["Matches"]] is not None else None
        goals = int(r[idx["Goals"]]) if r[idx["Goals"]] is not None else None
        ast = int(r[idx["Assists"]]) if r[idx["Assists"]] is not None else None
        mv = r[idx["Market_Value_EUR"]]
        mv_m = float(mv) / 1_000_000.0 if mv is not None else None

        extra = {}
        for col in hdr:
            if col in entity_cols:
                continue
            v = r[idx[col]]
            if v is None:
                continue
            key = col.replace("%", "_pct").replace(" ", "_")
            if isinstance(v, (int, float, str, bool)):
                extra[key] = v
            else:
                extra[key] = str(v)

        rec = {
            "playerExternalKey": pid,
            "fullName": name,
            "position": pos,
            "nationality": nat,
            "age": age,
            "appearances": apps,
            "goals": goals,
            "assists": ast,
            "marketValueMillionsEuro": mv_m,
            "eventTimestamp": None,
        }
        if extra:
            rec["extraFeatures"] = extra
        out.append(rec)
    wb.close()
    return out


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    matches = build_matches()
    players = build_players()
    with open(os.path.join(OUT_DIR, "ml_match_features_seed.json"), "w", encoding="utf-8") as f:
        json.dump(matches, f, indent=2, ensure_ascii=False)
    with open(os.path.join(OUT_DIR, "ml_player_features_seed.json"), "w", encoding="utf-8") as f:
        json.dump(players, f, indent=2, ensure_ascii=False)
    print(f"Wrote {len(matches)} match rows, {len(players)} player rows -> {OUT_DIR}")


if __name__ == "__main__":
    main()
